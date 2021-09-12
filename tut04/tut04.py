import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

 
def output_by_subject():
    FOLDER = "output_by_subject"
    header = ["rollno","register_sem","subno","sub_type"]
    sub_list = []
    
    if not os.path.exists(FOLDER): 
        os.makedirs(FOLDER)
    with open('regtable_old.csv','r') as file:
        stud_list=csv.reader(file)
        for data1 in stud_list:
            data= []
            data.append(data1[0])
            data.append(data1[1])
            data.append(data1[3])
            data.append(data1[8])
            if (data[2] =="subno"):continue
            if data[2] not in sub_list:
             sub_list.append(data[2])
             wb=Workbook()
             sheet=wb.active
             sheet.append(header)
             sheet.append(data)
             wb.save(f'output_by_subject\\{data[2]}.xlsx')
            else:
              wb=load_workbook(r'output_by_subject\\{}.xlsx'.format(data[2]))
              sheet=wb.active
              sheet.append(data)
              wb.save(f'output_by_subject\\{data[2]}.xlsx')
    return
 
def output_individual_roll():
    FOLDER = "output_individual_roll"
    header = ["rollno","register_sem","subno","sub_type"]
    roll_list = []
    
    if not os.path.exists(FOLDER): 
        os.makedirs(FOLDER)
    with open('regtable_old.csv','r') as file:
        stud_list=csv.reader(file)
        for data1 in stud_list:
            data= []
            data.append(data1[0])
            data.append(data1[1])
            data.append(data1[3])
            data.append(data1[8])
            if (data[0] =="rollno"):continue
            if data[0] not in roll_list:
             roll_list.append(data[0])
             wb=Workbook()
             sheet=wb.active
             sheet.append(header)
             sheet.append(data)
             wb.save(f'output_individual_roll\\{data[0]}.xlsx')
            else:
              wb=load_workbook(r'output_individual_roll\\{}.xlsx'.format(data[0]))
              sheet=wb.active
              sheet.append(data)
              wb.save(f'output_individual_roll\\{data[0]}.xlsx')
    return
 
output_by_subject()
output_individual_roll()
