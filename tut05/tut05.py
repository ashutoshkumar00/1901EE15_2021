def generate_marksheet():
    

# Import  Libraries
   import csv
   import os
   import shutil
   from openpyxl import Workbook
   from openpyxl import load_workbook
   from openpyxl.reader.excel import load_workbook
   os.system('cls')

    # Check if grades folder  exists, If yes delete and  remake
   if(os.path.isdir(r'.\output')):
       shutil.rmtree('.\output')
   os.makedirs('.\output')

   #using dictionary we map all 3 input file in dictionary course_data, name_data, roll_map
   roll_map={}
   course_data={}
   name_data={}
   list1=[]
    # grades map
   grades = {'AA': 10, 'AB': 9, 'BB': 8, 'BC': 7,
          'CC': 6, 'CD': 5, 'DD': 4,'DD*':4,'F*':0, 'F': 0, 'I': 0}  
   def overall_file_generator(a):
            wb = Workbook()
            wb.create_sheet(index=0,title="Overall")
            sheet=wb["Overall"]
            sheet.append(["Roll No.",a])
            sheet.append(["Name of Student ",name_data[a]])

            sheet.append(["Discipline",a[4:6]])
          
            no=[]
            spi_stud=[]
            credit_sem=[]
            total_credit=[]
            cpi_stud=[]

            for j in range(1,11):
               v=(a,str(j))
               if v in roll_map:
                     count=0
                     credit_obtain=0
                     for i in roll_map[v]:
                       count+=int(i[1])
                       credit_obtain+=grades[i[3]]*(int(i[1]))
               if v in roll_map:
                  no.append(j)   
                  credit_sem.append(count)     
                  spi_stud.append(credit_obtain/count)
            credit_sum=0
            sum=0
            for i in range(len(spi_stud)):
                   credit_sum+=credit_sem[i]
                   sum+=spi_stud[i]*credit_sem[i]
                   cpi_stud.append(sum/credit_sum)
                   total_credit.append(credit_sum)




            sheet.append(["Semester No."]+[  i for i in no])
            sheet.append(["Semester wise Credit Taken"]+[i for i in credit_sem])
            sheet.append(["SPI"]+[round(i,2) for i in spi_stud])
            sheet.append(["Total Credits Taken"]+[i for i in total_credit])
            sheet.append(["CPI"]+[round(i,2) for i in cpi_stud])
            wb.remove(wb["Sheet"])
            wb.save(".\output\{}.xlsx".format(a))  

 #Open the name_roll
   with open('names-roll.csv', 'r') as f:
      reader = csv.DictReader(f)
      for row in reader:
         name_data[row['Roll']]=row['Name']
         list1.append(row['Roll']) 

  # store grades in a roll_map     
   with open('grades.csv', 'r') as f:
      reader = csv.DictReader(f)
      for row in reader:
         x=(row['Roll'],row['Sem'])
         if x in roll_map:
               roll_map[x] +=[[row['SubCode'],row['Credit'],row['Sub_Type'],row['Grade'].strip()]]
         else:
             roll_map[x] =[[row['SubCode'],row['Credit'],row['Sub_Type'],row['Grade'].strip()]]    

   # store subject_master in  course_data
   with open('subjects_master.csv','r') as f:
      reader = csv.DictReader(f)
      for row in reader:
          course_data[row['subno']]=[row['subname'],row['ltp'],row['crd']]           

 #  overall sheet for all roll no
   for row in list1:
          overall_file_generator(row)

# add the semester detail  
   with open('names-roll.csv', 'r') as file:
      reader = csv.DictReader(file)
      for row in reader:
          a=row['Roll']       
          wb=load_workbook(".\output\{}.xlsx".format(a))
          p = 0
          for j in range(1,11):
               v=(a,str(j))
               if v in roll_map:
                    p=p+1
                    wb.create_sheet(index=p,title="Sem{}".format(str(j)))
                    wb["Sem{}".format(str(j))].append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
                    sheet=wb["Sem{}".format(str(j))]
                    for i in roll_map[v]:
                         x=course_data[i[0]]
                         sheet.append([ sheet.max_row,i[0],x[0],x[1],x[2],i[2],i[3] ])
          wb.save(".\output\{}.xlsx".format(a))

   return  
          
generate_marksheet()
     







