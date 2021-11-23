

# Import  Libraries
import csv
import os
import re
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
from openpyxl.reader.excel import load_workbook
os.system('cls')
   
def feedback_not_submitted():
   ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3:'practical'}
   output_file_name = "course_feedback_remaining.xlsx" 

   course_info={}
   student_info={}
   course_reg_info={}



   with open('course_master_dont_open_in_excel.csv','r') as file:
     reader = csv.DictReader(file)
     for row in reader:
	     course_info[row['subno']]=row['ltp']
		
   file.close()

   with open('studentinfo.csv','r') as f:
	   reader=csv.DictReader(f)
	   for row in reader:
		   student_info[row['Roll No']]=[row['Name'],row['email'],row['aemail'],int(row['contact'])]
   file.close()

   with open('course_registered_by_all_students.csv','r') as file:
	   reader=csv.DictReader(file)
	   for row in reader:
		   x=(row['rollno'],row['subno'])
		   course_reg_info[x]=[row['rollno'],row['subno'],int(row['register_sem']),int(row['schedule_sem']),0,0,0]
   file.close()

   with open('course_feedback_submitted_by_students.csv','r') as file:
	   reader=csv.DictReader(file)
	   for row in reader:
		   x=(row['stud_roll'],row['course_code'])
		   a=int(row['feedback_type'])
		   course_reg_info[x][a+3]=1
   file.close()   
   wb=openpyxl.load_workbook('./course_feedback_remaining.xlsx')
   sheet=wb['Sheet1']
   sheet.delete_rows(2, sheet.max_row+1)
   for row in course_reg_info:
       if row[1] not in course_info: continue
       l_t_p=course_info[row[1]]
       l=course_reg_info[row][4]	   
       t=course_reg_info[row][5]
       p=course_reg_info[row][6]
       lis=re.split("-",l_t_p)
       ll=lis[0]
       tt=lis[1]
       pp=lis[2]
       if((pp!="0" and p==0) or (tt!="0" and t==0) or (ll!="0" and l==0)):
		      if course_reg_info[row][0] in student_info:
		        sheet.append([course_reg_info[row][0],course_reg_info[row][2],course_reg_info[row][3],course_reg_info[row][1]]+student_info[course_reg_info[row][0]])
		      else:
		         sheet.append([course_reg_info[row][0],course_reg_info[row][2],course_reg_info[row][3],course_reg_info[row][1],'N/A','N/A','N/A','N/A'])
         
   wb.save('./course_feedback_remaining.xlsx')
feedback_not_submitted()

